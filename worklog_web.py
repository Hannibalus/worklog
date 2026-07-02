#!/usr/bin/env python3
#
# Work log - a small self-hosted work-logging app.
# Copyright (C) 2026 Andrej Lacho
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
from __future__ import annotations

import csv
import io
import os
import sqlite3
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional

from flask import Flask, g, jsonify, redirect, render_template_string, request, send_file, url_for, session, abort
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import secrets
from werkzeug.security import generate_password_hash, check_password_hash

APP = Flask(__name__)
DB_FILE = Path("/data/worklog.sqlite3")

# ----------------------------
# Config (via environment)
# ----------------------------
APP.secret_key = os.environ.get("WORKLOG_SECRET_KEY", "change-me-please-override-in-prod")

# Branding (override via env to customize)
BRAND_NAME = os.environ.get("WORKLOG_BRAND_NAME", "Work log")
BRAND_MARK = os.environ.get("WORKLOG_BRAND_MARK", "W")

# Bootstrap superadmin: created on first start if the users table is empty.
BOOTSTRAP_ADMIN_EMAIL = (os.environ.get("WORKLOG_ADMIN_EMAIL") or "").strip().lower()
BOOTSTRAP_ADMIN_PASSWORD = os.environ.get("WORKLOG_ADMIN_PASSWORD") or ""
BOOTSTRAP_ADMIN_NAME = os.environ.get("WORKLOG_ADMIN_NAME") or "Admin"

# Registration (invite) token lifetime, in hours.
INVITE_TTL_HOURS = int(os.environ.get("WORKLOG_INVITE_TTL_HOURS", "24"))  # 24 hours


# ----------------------------
# Helpers
# ----------------------------
def parse_date(s: str) -> date:
    s = (s or "").strip()
    return datetime.strptime(s, "%Y-%m-%d").date()

def fmt_date_ddmmyyyy(d: date) -> str:
    return d.strftime("%d-%m-%Y")

def parse_time(s: str) -> time:
    s = (s or "").strip()
    return datetime.strptime(s, "%H:%M").time()

def minutes_of(t: time) -> int:
    return t.hour * 60 + t.minute

def hours_between(d: date, t_from: time, t_to: time) -> float:
    dt_from = datetime.combine(d, t_from)
    dt_to = datetime.combine(d, t_to)
    if dt_to <= dt_from:
        raise ValueError("To musí byť väčšie ako From (blok cez polnoc je zakázaný).")
    minutes = (dt_to - dt_from).total_seconds() / 60.0
    return round((minutes / 60.0) * 10) / 10  # 0.1 h

def intervals_overlap(a_from: int, a_to: int, b_from: int, b_to: int) -> bool:
    return a_from < b_to and b_from < a_to

def format_work_text_for_display(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    if "\n" not in s and " - " in s:
        s = s.replace(" - ", "\n- ")
        if not s.lstrip().startswith("-"):
            s = s.replace("\n- ", "\n")
    return s

def md_escape(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("|", "\\|")
    s = s.replace("\n", "<br>")
    return s


# ----------------------------
# Auth/session
# ----------------------------
@dataclass
class UserCtx:
    id: int
    email: str
    name: str
    is_superadmin: bool

def current_user() -> Optional["UserCtx"]:
    uid = session.get("user_id")
    if not uid:
        return None
    row = get_db().execute(
        "SELECT id, email, name, is_superadmin, active FROM users WHERE id=?",
        (uid,),
    ).fetchone()
    if not row or not row["active"]:
        return None
    return UserCtx(
        id=int(row["id"]),
        email=row["email"],
        name=(row["name"] or row["email"]),
        is_superadmin=bool(row["is_superadmin"]),
    )

def require_login():
    if not current_user():
        return redirect(url_for("login", next=request.full_path))
    return None

def require_superadmin():
    u = current_user()
    if not u:
        return redirect(url_for("login", next=request.full_path))
    if not u.is_superadmin:
        abort(403)
    return None


# ----------------------------
# Users
# ----------------------------
def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def get_user_by_email(email: str):
    return get_db().execute(
        "SELECT * FROM users WHERE email=?", ((email or "").strip().lower(),)
    ).fetchone()

def get_user_by_id(uid: int):
    return get_db().execute("SELECT * FROM users WHERE id=?", (int(uid),)).fetchone()

def user_display(uid: int) -> str:
    r = get_user_by_id(uid)
    if not r:
        return f"#{uid}"
    return (r["name"] or r["email"])


# ----------------------------
# Companies / membership / roles
# ----------------------------
# Global role:   is_superadmin  -> can create companies, invite users, manage everything, see all.
# Per-company:   company_members.role in ('member','admin')
#                'member' -> logs own work, sees only own records
#                'admin'  -> read-only overseer: can view all records inside that one company
def companies_for_user(u: "UserCtx"):
    db = get_db()
    if u.is_superadmin:
        return db.execute("SELECT id, name FROM companies ORDER BY name").fetchall()
    return db.execute(
        "SELECT c.id, c.name FROM companies c "
        "JOIN company_members m ON m.company_id=c.id "
        "WHERE m.user_id=? ORDER BY c.name",
        (u.id,),
    ).fetchall()

def company_role(u: "UserCtx", company_id) -> Optional[str]:
    """Return 'admin' / 'member' / None (no access). Superadmin is always 'admin'."""
    if company_id is None:
        return None
    if u.is_superadmin:
        row = get_db().execute("SELECT id FROM companies WHERE id=?", (company_id,)).fetchone()
        return "admin" if row else None
    row = get_db().execute(
        "SELECT role FROM company_members WHERE user_id=? AND company_id=?",
        (u.id, company_id),
    ).fetchone()
    return row["role"] if row else None

def is_company_admin(u: "UserCtx", company_id) -> bool:
    return company_role(u, company_id) == "admin"

def current_company(u: "UserCtx"):
    """Return (company_id, role) for the currently selected company, or (None, None)."""
    cid = session.get("company_id")
    if cid is not None:
        role = company_role(u, cid)
        if role:
            return int(cid), role
    comps = companies_for_user(u)
    if comps:
        cid = int(comps[0]["id"])
        session["company_id"] = cid
        return cid, company_role(u, cid)
    return None, None

def effective_target_id(u: "UserCtx", company_id, target_id) -> int:
    """member -> always own id; company admin / superadmin -> may read another member's id."""
    if target_id and is_company_admin(u, company_id):
        try:
            tid = int(target_id)
        except Exception:
            return u.id
        row = get_db().execute(
            "SELECT 1 FROM company_members WHERE user_id=? AND company_id=?",
            (tid, company_id),
        ).fetchone()
        if row:
            return tid
    return u.id


# ----------------------------
# DB
# ----------------------------
def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON;")
        g.db = conn
    return g.db

@APP.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(DB_FILE)
    db.execute("PRAGMA foreign_keys=ON;")

    db.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL UNIQUE,
        name TEXT NOT NULL DEFAULT '',
        password_hash TEXT,
        is_superadmin INTEGER NOT NULL DEFAULT 0,
        active INTEGER NOT NULL DEFAULT 0,
        invite_token TEXT,
        invite_expires TEXT,
        created TEXT NOT NULL DEFAULT ''
    );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_users_email ON users(email);")
    db.execute("CREATE INDEX IF NOT EXISTS idx_users_token ON users(invite_token);")

    db.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        created TEXT NOT NULL DEFAULT ''
    );
    """)
    db.execute("""
    CREATE TABLE IF NOT EXISTS company_members (
        company_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        role TEXT NOT NULL DEFAULT 'member',
        PRIMARY KEY (company_id, user_id),
        FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    """)

    db.execute("""
    CREATE TABLE IF NOT EXISTS days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        company_id INTEGER NOT NULL,
        work_date TEXT NOT NULL,
        work_text TEXT NOT NULL DEFAULT '',
        UNIQUE(user_id, company_id, work_date)
    );
    """)
    db.execute("""
    CREATE TABLE IF NOT EXISTS blocks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        company_id INTEGER NOT NULL,
        day_id INTEGER NOT NULL,
        t_from TEXT NOT NULL,
        t_to TEXT NOT NULL,
        hours REAL NOT NULL,
        FOREIGN KEY(day_id) REFERENCES days(id) ON DELETE CASCADE
    );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_days_user_comp_date ON days(user_id, company_id, work_date);")
    db.execute("CREATE INDEX IF NOT EXISTS idx_blocks_user_comp_day ON blocks(user_id, company_id, day_id);")

    # Audit: superadmin / company-admin viewing or exporting another member's data.
    db.execute("""
    CREATE TABLE IF NOT EXISTS audit_views (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts_utc TEXT NOT NULL,
        company_id INTEGER NOT NULL DEFAULT 0,
        viewer_id INTEGER NOT NULL,
        target_id INTEGER NOT NULL,
        action TEXT NOT NULL,           -- month_view / export_csv / export_md / export_xlsx
        month TEXT NOT NULL DEFAULT '',
        ip TEXT NOT NULL DEFAULT '',
        user_agent TEXT NOT NULL DEFAULT ''
    );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_audit_ts ON audit_views(ts_utc DESC);")

    # Bootstrap superadmin (only if configured and the email is not present yet).
    if BOOTSTRAP_ADMIN_EMAIL and BOOTSTRAP_ADMIN_PASSWORD:
        ex = db.execute("SELECT id FROM users WHERE email=?", (BOOTSTRAP_ADMIN_EMAIL,)).fetchone()
        if not ex:
            db.execute(
                "INSERT INTO users(email, name, password_hash, is_superadmin, active, created) "
                "VALUES(?,?,?,1,1,?)",
                (BOOTSTRAP_ADMIN_EMAIL, BOOTSTRAP_ADMIN_NAME,
                 generate_password_hash(BOOTSTRAP_ADMIN_PASSWORD), now_iso()),
            )

    db.commit()
    db.close()

def audit_log(company_id, viewer_id: int, target_id: int, action: str, month: str = "") -> None:
    """Record a superadmin / company-admin view or export of another member's data."""
    try:
        db = get_db()
        ip = (request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
              or request.headers.get("X-Real-IP", "").strip()
              or request.remote_addr
              or "")
        ua = request.headers.get("User-Agent", "") or ""
        db.execute(
            "INSERT INTO audit_views(ts_utc, company_id, viewer_id, target_id, action, month, ip, user_agent) "
            "VALUES(?,?,?,?,?,?,?,?)",
            (now_iso(), int(company_id or 0), int(viewer_id), int(target_id), action, month or "", ip, ua),
        )
        db.commit()
    except Exception:
        # audit must never break the app
        pass

def ensure_day(d: date, user_id: int, company_id: int) -> int:
    db = get_db()
    # Make sure the writer is a member of this company (superadmins may log work
    # in any company without being an explicit member) so their records show up
    # for that company's managers. INSERT OR IGNORE keeps an existing role.
    db.execute(
        "INSERT OR IGNORE INTO company_members(company_id, user_id, role) VALUES(?,?, 'member')",
        (company_id, user_id),
    )
    db.execute(
        "INSERT OR IGNORE INTO days(user_id, company_id, work_date, work_text) VALUES(?,?,?, '')",
        (user_id, company_id, d.isoformat()),
    )
    db.commit()
    row = db.execute(
        "SELECT id FROM days WHERE user_id=? AND company_id=? AND work_date=?",
        (user_id, company_id, d.isoformat()),
    ).fetchone()
    return int(row["id"])

def get_day_row(d: date, user_id: int, company_id: int):
    return get_db().execute(
        "SELECT id, work_date, work_text FROM days WHERE user_id=? AND company_id=? AND work_date=?",
        (user_id, company_id, d.isoformat()),
    ).fetchone()

def fetch_blocks(day_id: int, user_id: int, company_id: int):
    return get_db().execute(
        "SELECT id, t_from, t_to, hours FROM blocks WHERE user_id=? AND company_id=? AND day_id=? ORDER BY t_from",
        (user_id, company_id, day_id),
    ).fetchall()

def fetch_blocks_intervals(day_id: int, user_id: int, company_id: int) -> List[Tuple[int, int]]:
    rows = get_db().execute(
        "SELECT t_from, t_to FROM blocks WHERE user_id=? AND company_id=? AND day_id=?",
        (user_id, company_id, day_id),
    ).fetchall()
    out: List[Tuple[int, int]] = []
    for r in rows:
        out.append((minutes_of(parse_time(r["t_from"])), minutes_of(parse_time(r["t_to"]))))
    return out

def day_total_hours(day_id: int, user_id: int, company_id: int) -> float:
    return float(get_db().execute(
        "SELECT COALESCE(SUM(hours),0) AS s FROM blocks WHERE user_id=? AND company_id=? AND day_id=?",
        (user_id, company_id, day_id),
    ).fetchone()["s"])

def month_total_hours(m: str, user_id: int, company_id: int) -> float:
    return float(get_db().execute(
        "SELECT COALESCE(SUM(b.hours),0) AS s "
        "FROM blocks b JOIN days d ON d.id=b.day_id "
        "WHERE b.user_id=? AND b.company_id=? AND substr(d.work_date,1,7)=?",
        (user_id, company_id, m),
    ).fetchone()["s"])

def rows_for_month(m: str, user_id: int, company_id: int) -> List[Dict[str, Any]]:
    db = get_db()
    days_rows = db.execute(
        "SELECT id, work_date, work_text FROM days WHERE user_id=? AND company_id=? AND substr(work_date,1,7)=? ORDER BY work_date",
        (user_id, company_id, m),
    ).fetchall()

    out: List[Dict[str, Any]] = []
    for drow in days_rows:
        did = int(drow["id"])
        blocks = db.execute(
            "SELECT t_from, t_to, hours FROM blocks WHERE user_id=? AND company_id=? AND day_id=? ORDER BY t_from",
            (user_id, company_id, did),
        ).fetchall()

        dtotal = day_total_hours(did, user_id, company_id)
        wtext = format_work_text_for_display(drow["work_text"])

        # one row per day even if it has no time blocks
        if not blocks:
            out.append({
                "day_id": did, "is_first": True, "work_date": drow["work_date"],
                "t_from": "", "t_to": "", "hours": "", "work_text": wtext,
                "day_total": f"{dtotal:.1f}",
            })
            continue

        first = True
        for b in blocks:
            out.append({
                "day_id": did, "is_first": first, "work_date": drow["work_date"],
                "t_from": b["t_from"], "t_to": b["t_to"], "hours": f"{float(b['hours']):.1f}",
                "work_text": wtext if first else "",
                "day_total": f"{dtotal:.1f}" if first else "",
            })
            first = False

    return out

def delete_day_by_id(day_id: int, user_id: int, company_id: int) -> None:
    db = get_db()
    db.execute("DELETE FROM days WHERE user_id=? AND company_id=? AND id=?", (user_id, company_id, int(day_id)))
    db.commit()

def company_users(company_id: int):
    """List of members (id, email, name, active, is_superadmin, role) in the given company."""
    return get_db().execute(
        "SELECT u.id, u.email, u.name, u.active, u.is_superadmin, m.role FROM company_members m "
        "JOIN users u ON u.id=m.user_id WHERE m.company_id=? ORDER BY u.name, u.email",
        (company_id,),
    ).fetchall()

def list_audit(limit: int = 500):
    return get_db().execute(
        "SELECT a.ts_utc, a.action, a.month, a.ip, "
        "vu.email AS viewer_email, tu.email AS target_email, c.name AS company_name "
        "FROM audit_views a "
        "LEFT JOIN users vu ON vu.id=a.viewer_id "
        "LEFT JOIN users tu ON tu.id=a.target_id "
        "LEFT JOIN companies c ON c.id=a.company_id "
        "ORDER BY a.id DESC LIMIT ?",
        (int(limit),),
    ).fetchall()


# ----------------------------
# UI templates
# ----------------------------
BASE_CSS = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root{
    /* SETUP.sk paleta */
    --bg:#f4f6f8;
    --bg-grad-1:#eef1f3;
    --bg-grad-2:#f7f9fa;
    --surface:#ffffff;
    --surface-2:#f8fafb;
    --fg:#2c3e50;
    --fg-soft:#3d4f5e;
    --muted:#7a8a99;
    --border:#e1e6ea;
    --border-strong:#cbd3da;
    --th:#eef2f4;
    --row-hover:#f6f9fa;
    /* Setup teal/cyan - primary action (presne setup.sk hex) */
    --accent:#4bbccb;
    --accent-hover:#0099b8;
    --accent-soft:#c7e7f0;
    --accent-fg:#ffffff;
    /* Setup yellow - sekundarna action (presne setup.sk hex) */
    --accent-2:#e9c809;
    --accent-2-hover:#d4b400;
    --accent-2-soft:#fdf5b4;
    --accent-2-fg:#3d3300;
    /* Dark topbar (setup.sk #38373a) */
    --topbar-bg:#38373a;
    --topbar-bg-2:#2c2b2e;
    --topbar-fg:#ffffff;
    --topbar-muted:#bdbcbe;
    --topbar-border:rgba(255,255,255,.08);
    --subnav-bg:#4bbccb;
    --subnav-fg:#ffffff;
    --danger:#fb4c49;
    --danger-soft:#ffe2e1;
    --success:#29a645;
    --success-soft:#dff5e3;
    --warning:#e9c809;
    --shadow-xs:0 1px 2px rgba(44,62,80,.05);
    --shadow-sm:0 1px 3px rgba(44,62,80,.07), 0 1px 2px -1px rgba(44,62,80,.04);
    --shadow:0 4px 12px -2px rgba(44,62,80,.07), 0 2px 6px -2px rgba(44,62,80,.04);
    --shadow-md:0 10px 20px -6px rgba(44,62,80,.10), 0 4px 10px -4px rgba(44,62,80,.05);
    --radius-sm:6px;
    --radius:10px;
    --radius-lg:14px;
  }
  .dark{
    --bg:#1a2330;
    --bg-grad-1:#15202d;
    --bg-grad-2:#1d2937;
    --surface:#243140;
    --surface-2:#2a3a4c;
    --fg:#e9eef2;
    --fg-soft:#c9d3dc;
    --muted:#8aa0b3;
    --border:#324456;
    --border-strong:#3f5468;
    --th:#1f2a37;
    --row-hover:#2a3a4c;
    --accent:#3bc6d4;
    --accent-hover:#5cd2de;
    --accent-soft:rgba(59,198,212,.14);
    --accent-fg:#0a1620;
    --accent-2:#fbb725;
    --accent-2-hover:#fcc24a;
    --accent-2-soft:rgba(233,200,9,.14);
    --accent-2-fg:#1a1100;
    --topbar-bg:#1f2a37;
    --topbar-bg-2:#1a242f;
    --topbar-fg:#e9eef2;
    --topbar-muted:#8aa0b3;
    --topbar-border:rgba(255,255,255,.06);
    --subnav-bg:#26a3b1;
    --subnav-fg:#ffffff;
    --danger:#ef5350;
    --danger-soft:rgba(239,83,80,.14);
    --success:#4caf50;
    --success-soft:rgba(76,175,80,.14);
    --warning:#ffa726;
    --shadow-xs:0 1px 2px rgba(0,0,0,.4);
    --shadow-sm:0 1px 3px rgba(0,0,0,.4), 0 1px 2px -1px rgba(0,0,0,.3);
    --shadow:0 4px 12px -2px rgba(0,0,0,.45), 0 2px 6px -2px rgba(0,0,0,.3);
    --shadow-md:0 10px 20px -6px rgba(0,0,0,.55), 0 4px 10px -4px rgba(0,0,0,.35);
  }

  *,*:before,*:after{ box-sizing:border-box; }
  html,body{ height:100%; }
  body{
    margin:0;
    font-family:'Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;
    font-size:14px;
    line-height:1.55;
    color:var(--fg);
    background:linear-gradient(180deg, var(--bg-grad-1), var(--bg-grad-2));
    background-attachment:fixed;
    -webkit-font-smoothing:antialiased;
    -moz-osx-font-smoothing:grayscale;
  }
  .dark body{
    background:linear-gradient(180deg, #11181f, #1a2330);
  }

  .container{ max-width:1280px; margin:0 auto; padding:24px; }

  h1,h2,h3,h4{ margin:0; font-weight:700; letter-spacing:-0.01em; color:var(--fg); }
  h2{ font-size:22px; }
  h3{ font-size:16px; }

  /* TOPBAR - setup.sk style: dark slate */
  .topbar{
    display:flex; align-items:center; gap:10px; flex-wrap:wrap;
    padding:10px 16px;
    background:linear-gradient(180deg, var(--topbar-bg), var(--topbar-bg-2));
    color:var(--topbar-fg);
    border:1px solid var(--topbar-bg-2);
    border-radius:var(--radius);
    box-shadow:var(--shadow-sm);
    margin-bottom:14px;
  }
  .topbar a, .topbar .btn{ color:var(--topbar-fg); }
  .topbar .btn{
    background:rgba(255,255,255,.05);
    border-color:var(--topbar-border);
    color:var(--topbar-fg);
    box-shadow:none;
  }
  .topbar .btn:hover{
    background:rgba(255,255,255,.12);
    border-color:rgba(255,255,255,.18);
    color:var(--accent-2);
  }
  .topbar .btn-primary{
    background:var(--accent);
    border-color:var(--accent);
    color:var(--accent-fg);
    box-shadow:0 4px 10px -3px rgba(75,188,203,.45);
  }
  .topbar .btn-primary:hover{
    background:var(--accent-hover);
    border-color:var(--accent-hover);
    color:var(--accent-fg);
  }
  .topbar .btn-ghost{ background:transparent; border-color:transparent; }
  .topbar .btn-ghost:hover{ background:rgba(255,255,255,.10); border-color:transparent; }
  .topbar input, .topbar select{
    background:rgba(255,255,255,.95);
    color:var(--fg);
    border-color:rgba(0,0,0,.05);
    width:auto;
  }
  .topbar input:focus, .topbar select:focus{
    border-color:var(--accent);
    box-shadow:0 0 0 3px rgba(75,188,203,.30);
  }
  .topbar .small{ color:var(--topbar-muted); }

  .brand{ display:flex; align-items:center; gap:10px; margin-right:6px; }
  .brand-mark{
    display:inline-flex; align-items:center; justify-content:center;
    width:34px; height:34px; border-radius:8px; color:#fff; font-weight:800; font-size:15px;
    background:linear-gradient(135deg, var(--accent) 0%, var(--accent-hover) 100%);
    box-shadow:0 4px 10px -2px rgba(75,188,203,.55), inset 0 1px 0 rgba(255,255,255,.25);
    letter-spacing:-0.02em;
  }
  .brand-title{ font-weight:700; font-size:15px; letter-spacing:-0.01em; color:var(--topbar-fg); }
  .brand-sub{ font-size:11px; color:var(--topbar-muted); margin-top:-2px; }
  .topbar .spacer{ flex:1; }
  .userchip{
    display:inline-flex; align-items:center; gap:8px;
    padding:5px 12px 5px 5px;
    background:rgba(255,255,255,.08);
    border:1px solid var(--topbar-border);
    border-radius:999px;
    font-size:12px; color:var(--topbar-fg);
  }
  .userchip .avatar{
    width:26px; height:26px; border-radius:50%;
    background:linear-gradient(135deg, var(--accent), var(--accent-hover));
    color:#fff; display:inline-flex; align-items:center; justify-content:center;
    font-weight:700; font-size:11px; text-transform:uppercase;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.2);
  }
  .userchip .role{
    font-size:10px; padding:2px 7px; border-radius:999px;
    background:var(--accent-2); color:var(--accent-2-fg);
    font-weight:700; letter-spacing:.04em; text-transform:uppercase;
  }

  /* LAYOUT */
  .row{ display:grid; grid-template-columns:1fr 1fr; gap:18px; align-items:start; }
  @media (max-width:980px){ .row{ grid-template-columns:1fr; } }

  .stat-row{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; margin-bottom:18px; }
  @media (max-width:760px){ .stat-row{ grid-template-columns:1fr; } }
  .stat{
    background:var(--surface);
    border:1px solid var(--border);
    border-radius:var(--radius);
    padding:16px 18px;
    box-shadow:var(--shadow-sm);
    position:relative; overflow:hidden;
  }
  .stat:before{
    content:""; position:absolute; inset:0 auto 0 0; width:4px;
    background:linear-gradient(180deg, var(--accent), var(--accent-hover));
  }
  .stat.stat-2:before{ background:linear-gradient(180deg, var(--accent-2), var(--accent-2-hover)); }
  .stat .value{ color:var(--fg); }
  .stat .label{ font-size:11px; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; font-weight:600; }
  .stat .value{ font-size:28px; font-weight:700; letter-spacing:-0.02em; margin-top:4px; }
  .stat .value small{ font-size:13px; font-weight:500; color:var(--muted); margin-left:4px; }
  .stat .hint{ font-size:12px; color:var(--muted); margin-top:2px; }

  /* CARD */
  .card{
    background:var(--surface);
    border:1px solid var(--border);
    border-radius:var(--radius);
    padding:18px;
    box-shadow:var(--shadow-sm);
    transition:box-shadow .2s ease, border-color .2s ease;
  }
  .card-title{
    display:flex; align-items:center; gap:8px;
    font-size:15px; font-weight:600; color:var(--fg);
    margin:-2px 0 14px 0;
  }
  .card-title .ico{
    width:30px; height:30px; border-radius:8px;
    background:var(--accent-soft); color:var(--accent);
    display:inline-flex; align-items:center; justify-content:center;
    border:1px solid rgba(75,188,203,.18);
  }
  .card-title .ico svg{ width:16px; height:16px; }

  /* ACTIONS */
  .actions{ display:flex; gap:10px; flex-wrap:wrap; align-items:end; }
  .spacer{ flex:1; }
  .small{ color:var(--muted); font-size:11px; font-weight:500; text-transform:uppercase; letter-spacing:.06em; }
  .muted{ color:var(--muted); }

  /* FIELDS */
  .field{ display:flex; flex-direction:column; gap:5px; }
  input, select, textarea{
    font-family:inherit;
    font-size:14px;
    padding:9px 12px;
    box-sizing:border-box;
    width:100%;
    background:var(--surface);
    color:var(--fg);
    border:1px solid var(--border-strong);
    border-radius:var(--radius-sm);
    transition:border-color .15s ease, box-shadow .15s ease, background .15s ease;
    outline:none;
  }
  input:focus, select:focus, textarea:focus{
    border-color:var(--accent);
    box-shadow:0 0 0 3px var(--accent-soft);
  }
  input[type="date"], input[type="time"]{ font-variant-numeric:tabular-nums; }
  textarea{
    font-family:'JetBrains Mono', ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;
    min-height:340px;
    line-height:1.6;
    resize:vertical;
  }
  input:disabled, select:disabled, textarea:disabled{
    background:var(--surface-2); color:var(--muted); cursor:not-allowed;
  }

  /* BUTTONS */
  button, .btn{
    display:inline-flex; align-items:center; justify-content:center; gap:6px;
    padding:9px 14px;
    border:1px solid var(--border-strong);
    background:var(--surface);
    border-radius:var(--radius-sm);
    cursor:pointer;
    color:var(--fg);
    text-decoration:none;
    font-family:inherit;
    font-size:13px; font-weight:500;
    line-height:1.2;
    transition:transform .12s ease, box-shadow .15s ease, background .15s ease, border-color .15s ease, color .15s ease;
    box-shadow:var(--shadow-xs);
    white-space:nowrap;
  }
  button:hover:not(:disabled), .btn:hover{
    background:var(--row-hover);
    border-color:var(--accent);
    color:var(--accent);
    transform:translateY(-1px);
    box-shadow:var(--shadow-sm);
  }
  button:active, .btn:active{ transform:translateY(0); }
  button:disabled{ opacity:.55; cursor:not-allowed; }

  .btn-primary{
    background:var(--accent);
    color:var(--accent-fg);
    border-color:var(--accent);
    box-shadow:0 5px 12px -3px rgba(75,188,203,.40);
  }
  .btn-primary:hover:not(:disabled){
    background:var(--accent-hover);
    border-color:var(--accent-hover);
    color:var(--accent-fg);
    box-shadow:0 7px 16px -5px rgba(75,188,203,.55);
  }

  .btn-secondary{
    background:var(--accent-2);
    color:var(--accent-2-fg);
    border-color:var(--accent-2);
    box-shadow:0 5px 12px -3px rgba(233,200,9,.40);
    font-weight:600;
  }
  .btn-secondary:hover:not(:disabled){
    background:var(--accent-2-hover);
    border-color:var(--accent-2-hover);
    color:var(--accent-2-fg);
    box-shadow:0 7px 16px -5px rgba(233,200,9,.55);
  }

  .btn-ghost{
    background:transparent;
    border-color:transparent;
    box-shadow:none;
  }
  .btn-ghost:hover{ background:var(--row-hover); border-color:var(--border); }

  .autosave-status{
    font-size:12px;
    color:var(--muted);
    align-self:center;
    min-width:140px;
    transition:opacity .25s;
    opacity:0;
  }
  .autosave-status.visible{ opacity:1; }
  .autosave-status.saving{ color:var(--muted); }
  .autosave-status.saved{ color:#0a8a4a; }
  .autosave-status.error{ color:var(--danger); }

  .btn-danger{
    background:transparent;
    border-color:var(--border-strong);
    color:var(--danger);
  }
  .btn-danger:hover:not(:disabled){
    background:var(--danger-soft);
    border-color:var(--danger);
    color:var(--danger);
  }

  .btn-icon{
    width:38px; height:38px; padding:0;
  }
  .btn-icon svg{ width:18px; height:18px; }

  .preset-grp{ display:flex; gap:8px; flex-wrap:wrap; }
  .preset-grp form{ margin:0; }
  .preset-grp button{
    padding:7px 11px;
    background:var(--surface-2);
    border-style:dashed;
    color:var(--fg-soft);
    font-variant-numeric:tabular-nums;
    font-size:12.5px;
  }
  .preset-grp button:hover:not(:disabled){
    background:var(--accent-soft);
    color:var(--accent);
    border-style:solid;
  }

  /* TABLE */
  .table-wrap{
    border:1px solid var(--border);
    border-radius:var(--radius-sm);
    overflow:hidden;
    background:var(--surface);
  }
  table{ border-collapse:separate; border-spacing:0; width:100%; }
  th, td{ padding:10px 12px; vertical-align:top; text-align:left; }
  th{
    font-size:11px; font-weight:600; text-transform:uppercase; letter-spacing:.06em;
    color:var(--muted);
    background:var(--th);
    border-bottom:1px solid var(--border);
  }
  td{ font-size:13.5px; border-bottom:1px solid var(--border); color:var(--fg-soft); }
  tbody tr:last-child td{ border-bottom:none; }
  tbody tr{ transition:background .12s ease; }
  tbody tr:hover{ background:var(--row-hover); }
  td.num, th.num{ font-variant-numeric:tabular-nums; }

  .total{ font-weight:700; font-size:14px; color:var(--fg); }
  .work{ white-space:pre-wrap; line-height:1.45; color:var(--fg-soft); }

  .msg{
    margin-top:12px;
    padding:10px 14px;
    background:var(--danger-soft);
    border:1px solid var(--danger);
    color:var(--danger);
    border-radius:var(--radius-sm);
    font-size:13px; font-weight:500;
  }

  /* MISC */
  a{ color:var(--accent); }
  a:hover{ color:var(--accent-hover); }
  hr.sep{ border:none; border-top:1px solid var(--border); margin:14px 0; }

  /* THEME ICON */
  .theme-icon-sun, .theme-icon-moon{ display:none; }
  :root:not(.dark) .theme-icon-moon{ display:block; }
  .dark .theme-icon-sun{ display:block; }

  /* Scrollbar - subtle */
  ::-webkit-scrollbar{ width:10px; height:10px; }
  ::-webkit-scrollbar-track{ background:transparent; }
  ::-webkit-scrollbar-thumb{ background:var(--border-strong); border-radius:10px; border:2px solid transparent; background-clip:padding-box; }
  ::-webkit-scrollbar-thumb:hover{ background:var(--muted); background-clip:padding-box; border:2px solid transparent; }
</style>

<script>
(function(){
  function setTheme(mode){
    const root = document.documentElement;
    if(mode === 'dark') root.classList.add('dark');
    else root.classList.remove('dark');
    try{ localStorage.setItem('theme', mode); }catch(e){}
    document.querySelectorAll('#themeToggle').forEach(function(btn){
      btn.setAttribute('aria-label', mode === 'dark' ? 'Light mode' : 'Dark mode');
      btn.dataset.mode = mode;
    });
  }
  function getTheme(){
    try{
      const saved = localStorage.getItem('theme');
      if(saved === 'dark' || saved === 'light') return saved;
    }catch(e){}
    return (window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches) ? 'dark' : 'light';
  }
  // Apply theme as early as possible to avoid flash
  setTheme(getTheme());
  document.addEventListener('DOMContentLoaded', function(){
    setTheme(getTheme());
    document.querySelectorAll('#themeToggle').forEach(function(btn){
      btn.addEventListener('click', function(){
        const cur = document.documentElement.classList.contains('dark') ? 'dark' : 'light';
        setTheme(cur === 'dark' ? 'light' : 'dark');
      });
    });
  });
})();
</script>
"""

THEME_BTN = """
<button id="themeToggle" type="button" class="btn btn-icon btn-ghost" title="Prepnúť motív" aria-label="Theme">
  <svg class="theme-icon-moon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
    <path d="M21 12.79A9 9 0 1 1 11.21 3a7 7 0 0 0 9.79 9.79z"></path>
  </svg>
  <svg class="theme-icon-sun" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
    <circle cx="12" cy="12" r="4"></circle>
    <path d="M12 2v2M12 20v2M4.93 4.93l1.41 1.41M17.66 17.66l1.41 1.41M2 12h2M20 12h2M4.93 19.07l1.41-1.41M17.66 6.34l1.41-1.41"></path>
  </svg>
</button>
"""

BRAND_HTML = """
<div class="brand">
  <span class="brand-mark">{{ brand_mark }}</span>
  <div>
    <div class="brand-title">Work log</div>
    <div class="brand-sub">{{ brand_name }}</div>
  </div>
</div>
"""

LOGIN_TEMPLATE = """
<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Sign in – Work log</title>
  {{ css|safe }}
  <style>
    .login-shell{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:24px; }
    .login-card{ width:min(440px, 100%); }
    .login-hero{
      display:flex; flex-direction:column; align-items:center; gap:14px;
      margin-bottom:22px;
    }
    .login-logo{
      width:64px; height:64px; border-radius:14px;
      background:linear-gradient(135deg, var(--accent) 0%, var(--accent-hover) 100%);
      color:#fff; display:inline-flex; align-items:center; justify-content:center;
      font-weight:800; font-size:26px; letter-spacing:-0.02em;
      box-shadow:0 14px 30px -8px rgba(75,188,203,.50), inset 0 1px 0 rgba(255,255,255,.25);
    }
    .login-h{ font-size:24px; font-weight:700; letter-spacing:-0.02em; }
    .login-sub{ color:var(--muted); font-size:13px; margin-top:-4px; }
    .login-fields{ display:flex; flex-direction:column; gap:14px; }
    .field-label{ font-size:12px; font-weight:600; color:var(--fg-soft); margin-bottom:6px; }
    .login-foot{ text-align:center; font-size:11px; color:var(--muted); margin-top:18px; }
  </style>
</head>

<body>
  <div class="login-shell">
    <div class="login-card">
      <div style="display:flex; justify-content:flex-end; margin-bottom:10px;">
        {{ theme_btn|safe }}
      </div>

      <div class="login-hero">
        <div class="login-logo">{{ brand_mark }}</div>
        <div style="text-align:center;">
          <div class="login-h">Work log</div>
          <div class="login-sub">{{ brand_name }}</div>
        </div>
      </div>

      <div class="card" style="padding:22px;">
        {% if msg %}
          <div class="msg" style="margin-top:0; margin-bottom:14px;">{{ msg }}</div>
        {% endif %}

        <form method="post" class="login-fields">
          <div>
            <div class="field-label">Email</div>
            <input name="email" type="email" autocomplete="username" placeholder="you@example.com" required autofocus>
          </div>

          <div>
            <div class="field-label">Password</div>
            <input name="pw" type="password" autocomplete="current-password" placeholder="••••••••" required>
          </div>

          <button class="btn-primary" style="width:100%; padding:11px 14px; font-size:14px;">
            <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M15 3h4a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2h-4"></path><polyline points="10 17 15 12 10 7"></polyline><line x1="15" y1="12" x2="3" y2="12"></line></svg>
            Sign in
          </button>
        </form>
      </div>

      <div class="login-foot">Sign in with your email and password. Accounts are created by invitation.</div>
    </div>
  </div>
</body>
</html>
"""

DAILY_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Work log – {{ d_human }}</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">

    <!-- TOPBAR -->
    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div>
          <div class="brand-title">Work log</div>
          <div class="brand-sub">daily log</div>
        </div>
      </div>

      <form method="get" action="/" style="margin:0; display:flex; align-items:end; gap:8px;">
        <input type="date" name="d" value="{{ d_iso }}" style="width:auto;">
        <button class="btn-primary" type="submit">
          <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>
          Load
        </button>
      </form>

      <a class="btn" href="{{ url_for('today') }}" title="Today">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><polyline points="12 6 12 12 16 14"></polyline></svg>
        Today
      </a>

      {% if is_superadmin %}
      <a class="btn" href="{{ url_for('admin_companies') }}" title="Admin">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"></path></svg>
        Admin
      </a>
      {% endif %}

      <a class="btn" href="{{ url_for('month_view') }}">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="4" width="18" height="18" rx="2"></rect><line x1="16" y1="2" x2="16" y2="6"></line><line x1="8" y1="2" x2="8" y2="6"></line><line x1="3" y1="10" x2="21" y2="10"></line></svg>
        Month
      </a>

      <div class="spacer"></div>

      {% if companies and companies|length > 0 %}
      <form method="post" action="{{ url_for('switch_company') }}" style="margin:0; display:flex; align-items:end;" title="Switch company">
        <select name="company_id" onchange="this.form.submit()" style="width:auto;">
          {% for c in companies %}
            <option value="{{ c.id }}" {% if c.id == company_id %}selected{% endif %}>{{ c.name }}</option>
          {% endfor %}
        </select>
        <noscript><button class="btn" type="submit">Switch</button></noscript>
      </form>
      {% endif %}

      <div class="userchip" title="Signed in user">
        <span class="avatar">{{ user_name[:1]|upper }}</span>
        <span>{{ user_name }}</span>
        <span class="role">{{ role_label }}</span>
      </div>

      <a class="btn" href="{{ url_for('account') }}" title="Change password">Account</a>

      {{ theme_btn|safe }}

      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>

    {% if msg %}<div class="msg">{{ msg }}</div>{% endif %}

    <!-- STATS -->
    <div class="stat-row">
      <div class="stat">
        <div class="label">Date</div>
        <div class="value" style="font-size:22px;">{{ d_human }}</div>
        <div class="hint">{{ d_iso }}</div>
      </div>
      <div class="stat stat-2">
        <div class="label">Total for day</div>
        <div class="value">{{ "%.1f"|format(day_total) }}<small>h</small></div>
        <div class="hint">{{ blocks|length }} block{{ '' if blocks|length == 1 else 's' }}</div>
      </div>
      <div class="stat">
        <div class="label">Total for month {{ m }}</div>
        <div class="value">{{ "%.1f"|format(month_total) }}<small>h</small></div>
        <div class="hint">Monthly summary</div>
      </div>
    </div>

    <!-- MAIN -->
    <div class="row">

      <div class="card">
        <div class="card-title">
          <span class="ico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline><line x1="16" y1="13" x2="8" y2="13"></line><line x1="16" y1="17" x2="8" y2="17"></line></svg></span>
          Work description – {{ d_human }}
        </div>

        <form method="post" action="{{ url_for('save_day') }}" id="save-day-form">
          <input type="hidden" name="work_date" value="{{ d_iso }}">
          <textarea name="work_text" id="work-text" placeholder="- what did you work on..." {% if readonly %}disabled{% endif %}>{{ work_text }}</textarea>
          <div class="actions" style="margin-top:12px; flex-wrap:nowrap;">
            <button type="submit" class="btn-primary" {% if readonly %}disabled{% endif %}>
              <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"></path><polyline points="17 21 17 13 7 13 7 21"></polyline><polyline points="7 3 7 8 15 8"></polyline></svg>
              Save text
            </button>

            <span id="autosave-status" class="autosave-status" aria-live="polite"></span>

            <div style="display:flex; gap:8px; flex-wrap:nowrap; margin-left:auto;">
            <form method="post" action="{{ url_for('copy_yesterday_text') }}" style="margin:0;">
              <input type="hidden" name="work_date" value="{{ d_iso }}">
              <button type="submit" class="btn-secondary" {% if readonly %}disabled{% endif %} title="Copy yesterday's text">
                <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="9" y="9" width="13" height="13" rx="2"></rect><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path></svg>
                Yesterday text
              </button>
            </form>

            <form method="post" action="{{ url_for('copy_yesterday_blocks') }}" style="margin:0;">
              <input type="hidden" name="work_date" value="{{ d_iso }}">
              <button type="submit" class="btn-secondary" {% if readonly %}disabled{% endif %} title="Copy yesterday's time blocks">
                <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><polyline points="12 6 12 12 16 14"></polyline></svg>
                Yesterday blocks
              </button>
            </form>
            </div>
          </div>
        </form>
      </div>

      <div class="card">
        <div class="card-title">
          <span class="ico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><polyline points="12 6 12 12 16 14"></polyline></svg></span>
          Time blocks
        </div>

        <form method="post" action="{{ url_for('add_block') }}">
          <input type="hidden" name="work_date" value="{{ d_iso }}">
          <div class="actions" style="align-items:end;">
            <div class="field" style="flex:1;">
              <div class="small">From</div>
              <input type="time" name="t_from" value="{{ default_from }}" step="60" {% if readonly %}disabled{% endif %}>
            </div>
            <div class="field" style="flex:1;">
              <div class="small">To</div>
              <input type="time" name="t_to" value="{{ default_to }}" step="60" {% if readonly %}disabled{% endif %}>
            </div>
            <button class="btn-primary" {% if readonly %}disabled{% endif %}>
              <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="5" x2="12" y2="19"></line><line x1="5" y1="12" x2="19" y2="12"></line></svg>
              Add
            </button>
          </div>
        </form>

        <div class="small" style="margin-top:14px; margin-bottom:6px;">Quick presets</div>
        <div class="preset-grp">
          <form method="post" action="{{ url_for('add_preset') }}">
            <input type="hidden" name="work_date" value="{{ d_iso }}">
            <input type="hidden" name="t_from" value="08:00">
            <input type="hidden" name="t_to" value="15:30">
            <button {% if readonly %}disabled{% endif %}>08:00 – 15:30 <span style="opacity:.6;">(7.5h)</span></button>
          </form>
          <form method="post" action="{{ url_for('add_preset') }}">
            <input type="hidden" name="work_date" value="{{ d_iso }}">
            <input type="hidden" name="t_from" value="08:00">
            <input type="hidden" name="t_to" value="16:00">
            <button {% if readonly %}disabled{% endif %}>08:00 – 16:00 <span style="opacity:.6;">(8h)</span></button>
          </form>
          <form method="post" action="{{ url_for('add_preset') }}">
            <input type="hidden" name="work_date" value="{{ d_iso }}">
            <input type="hidden" name="t_from" value="09:00">
            <input type="hidden" name="t_to" value="17:00">
            <button {% if readonly %}disabled{% endif %}>09:00 – 17:00 <span style="opacity:.6;">(8h)</span></button>
          </form>
        </div>

        <div class="table-wrap" style="margin-top:16px;">
          <table>
            <thead>
              <tr>
                <th class="num">From</th>
                <th class="num">To</th>
                <th class="num">Hours</th>
                <th style="width:1%;"></th>
              </tr>
            </thead>
            <tbody>
              {% for b in blocks %}
              <tr>
                <td class="num">{{ b.t_from }}</td>
                <td class="num">{{ b.t_to }}</td>
                <td class="num"><b>{{ "%.1f"|format(b.hours) }}</b></td>
                <td>
                  <form method="post" action="{{ url_for('delete_block') }}" style="margin:0;">
                    <input type="hidden" name="work_date" value="{{ d_iso }}">
                    <input type="hidden" name="block_id" value="{{ b.id }}">
                    <button class="btn-danger btn-icon" title="Delete block" {% if readonly %}disabled{% endif %}>
                      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6l-2 14a2 2 0 0 1-2 2H9a2 2 0 0 1-2-2L5 6"></path><path d="M10 11v6M14 11v6"></path><path d="M9 6V4a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v2"></path></svg>
                    </button>
                  </form>
                </td>
              </tr>
              {% endfor %}
              {% if not blocks %}
              <tr><td colspan="4" class="muted" style="text-align:center; padding:24px;">No blocks for this day. Add your first →</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
  </div>

  {% if not readonly %}
  <script>
  (function(){
    const form    = document.getElementById('save-day-form');
    const ta      = document.getElementById('work-text');
    const status  = document.getElementById('autosave-status');
    if (!form || !ta || !status) return;

    const DEBOUNCE_MS = 600;
    let timer = null;
    let inFlight = null;
    let lastSaved = ta.value;
    let pending = false;

    function setStatus(state, text){
      status.className = 'autosave-status visible ' + state;
      status.textContent = text;
      if (state === 'saved') {
        setTimeout(() => {
          if (status.classList.contains('saved')) {
            status.classList.remove('visible');
          }
        }, 2000);
      }
    }

    async function doSave(){
      if (ta.value === lastSaved && !pending) return;
      pending = false;
      const value = ta.value;
      const fd = new FormData(form);
      setStatus('saving', 'Saving…');
      try {
        if (inFlight) { try { inFlight.abort(); } catch(e){} }
        const ctrl = new AbortController();
        inFlight = ctrl;
        const resp = await fetch(form.action, {
          method: 'POST',
          body: fd,
          headers: { 'X-Requested-With': 'fetch' },
          credentials: 'same-origin',
          signal: ctrl.signal,
        });
        inFlight = null;
        if (resp.status === 401) {
          setStatus('error', 'Signed out – reload the page');
          return;
        }
        if (!resp.ok) throw new Error('HTTP ' + resp.status);
        const data = await resp.json();
        lastSaved = value;
        setStatus('saved', 'Saved ' + (data.saved_at || ''));
      } catch(err) {
        if (err.name === 'AbortError') return;
        setStatus('error', 'Save failed – retrying');
        // retry once after 3s
        setTimeout(() => { pending = true; doSave(); }, 3000);
      }
    }

    function schedule(){
      pending = true;
      if (timer) clearTimeout(timer);
      timer = setTimeout(() => { timer = null; doSave(); }, DEBOUNCE_MS);
    }

    ta.addEventListener('input', schedule);
    ta.addEventListener('blur', () => {
      if (timer || pending) {
        if (timer) { clearTimeout(timer); timer = null; }
        doSave();
      }
    });

    form.addEventListener('submit', (e) => {
      e.preventDefault();
      if (timer) { clearTimeout(timer); timer = null; }
      pending = true;
      doSave();
    });

    window.addEventListener('beforeunload', (e) => {
      if (timer || pending || inFlight || ta.value !== lastSaved) {
        e.preventDefault();
        e.returnValue = '';
      }
    });
  })();
  </script>
  {% endif %}
</body></html>
"""

MONTH_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Monthly overview – {{ sel }}</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">

    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div>
          <div class="brand-title">Work log</div>
          <div class="brand-sub">monthly overview</div>
        </div>
      </div>

      <a class="btn" href="{{ url_for('daily') }}">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="19" y1="12" x2="5" y2="12"></line><polyline points="12 19 5 12 12 5"></polyline></svg>
        Back
      </a>

      <form method="get" action="{{ url_for('month_view') }}" style="margin:0; display:flex; align-items:end; gap:8px;">
        {% if can_view_others %}
        <select name="u" style="width:auto;">
          {% for mem in members %}
            <option value="{{ mem.id }}" {% if mem.id == target_id %}selected{% endif %}>{{ mem.name or mem.email }}</option>
          {% endfor %}
        </select>
        {% endif %}
        <select name="m" style="width:auto;">
          {% for mm in months %}
            <option value="{{ mm }}" {% if mm == sel %}selected{% endif %}>{{ mm }}</option>
          {% endfor %}
        </select>
        <button class="btn-primary">Show</button>
      </form>

      <div class="spacer"></div>

      {% if companies and companies|length > 0 %}
      <form method="post" action="{{ url_for('switch_company') }}" style="margin:0; display:flex; align-items:end;" title="Switch company">
        <select name="company_id" onchange="this.form.submit()" style="width:auto;">
          {% for c in companies %}
            <option value="{{ c.id }}" {% if c.id == company_id %}selected{% endif %}>{{ c.name }}</option>
          {% endfor %}
        </select>
        <noscript><button class="btn" type="submit">Switch</button></noscript>
      </form>
      {% endif %}

      <div class="userchip" title="Viewed user">
        <span class="avatar">{{ target_name[:1]|upper }}</span>
        <span>{{ target_name }}</span>
        {% if is_viewing_other %}<span class="role">read-only</span>{% endif %}
      </div>

      <a class="btn" href="{{ url_for('export_csv', m=sel, u=target_id) }}" title="Export CSV">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>
        CSV
      </a>
      <a class="btn" href="{{ url_for('export_md', m=sel, u=target_id) }}" title="Export Markdown">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>
        MD
      </a>
      <a class="btn" href="{{ url_for('export_xlsx', m=sel, u=target_id) }}" title="Export XLSX">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>
        XLSX
      </a>

      <a class="btn" href="{{ url_for('account') }}" title="Change password">Account</a>

      {{ theme_btn|safe }}

      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>

    <!-- STATS -->
    <div class="stat-row">
      <div class="stat">
        <div class="label">Month</div>
        <div class="value" style="font-size:24px;">{{ sel }}</div>
        <div class="hint">{{ target_name }}</div>
      </div>
      <div class="stat stat-2">
        <div class="label">Total</div>
        <div class="value">{{ "%.1f"|format(month_total) }}<small>h</small></div>
        <div class="hint">All blocks in the month</div>
      </div>
      <div class="stat">
        <div class="label">Days</div>
        {% set day_ids = [] %}{% for r in rows %}{% if r.is_first %}{% set _ = day_ids.append(r.day_id) %}{% endif %}{% endfor %}
        <div class="value">{{ day_ids|length }}<small>days</small></div>
        <div class="hint">Active work days</div>
      </div>
    </div>

    <div class="card" style="padding:0; overflow:hidden;">
      <div class="table-wrap" style="border:none; border-radius:0;">
        <table>
          <thead>
            <tr>
              <th style="width:110px;">Date</th>
              <th class="num" style="width:70px;">From</th>
              <th class="num" style="width:70px;">To</th>
              <th class="num" style="width:80px;">Hours</th>
              <th>Work</th>
              <th class="num" style="width:120px;">Day total</th>
              {% if can_delete_day %}<th style="width:1%;"></th>{% endif %}
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            <tr>
              <td><b style="font-variant-numeric:tabular-nums;">{{ r.work_date }}</b></td>
              <td class="num">{{ r.t_from }}</td>
              <td class="num">{{ r.t_to }}</td>
              <td class="num">{{ r.hours }}</td>
              <td><div class="work">{{ r.work_text }}</div></td>
              <td class="num"><b>{{ r.day_total }}</b></td>
              {% if can_delete_day %}
              <td>
                {% if r.is_first %}
                  <form method="post" action="{{ url_for('delete_day') }}" style="margin:0;">
                    <input type="hidden" name="m" value="{{ sel }}">
                    <input type="hidden" name="day_id" value="{{ r.day_id }}">
                    <button class="btn-danger btn-icon" title="Delete whole day">
                      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6l-2 14a2 2 0 0 1-2 2H9a2 2 0 0 1-2-2L5 6"></path></svg>
                    </button>
                  </form>
                {% endif %}
              </td>
              {% endif %}
            </tr>
            {% endfor %}
            {% if not rows %}
            <tr><td colspan="{{ 7 if can_delete_day else 6 }}" class="muted" style="text-align:center; padding:32px;">No data for this month.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <div style="padding:14px 18px; display:flex; align-items:center; justify-content:space-between; border-top:1px solid var(--border); background:var(--surface-2);">
        <div class="small">Monthly summary</div>
        <div class="total" style="font-size:18px;">{{ "%.1f"|format(month_total) }} h</div>
      </div>
    </div>
  </div>
</body></html>
"""

REGISTER_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Create account – Work log</title>
  {{ css|safe }}
  <style>
    .login-shell{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:24px; }
    .login-card{ width:min(460px, 100%); }
    .login-hero{ display:flex; flex-direction:column; align-items:center; gap:14px; margin-bottom:22px; }
    .login-logo{ width:64px; height:64px; border-radius:14px; background:linear-gradient(135deg, var(--accent) 0%, var(--accent-hover) 100%); color:#fff; display:inline-flex; align-items:center; justify-content:center; font-weight:800; font-size:26px; letter-spacing:-0.02em; box-shadow:0 14px 30px -8px rgba(75,188,203,.50), inset 0 1px 0 rgba(255,255,255,.25); }
    .login-h{ font-size:24px; font-weight:700; letter-spacing:-0.02em; }
    .login-sub{ color:var(--muted); font-size:13px; margin-top:-4px; }
    .login-fields{ display:flex; flex-direction:column; gap:14px; }
    .field-label{ font-size:12px; font-weight:600; color:var(--fg-soft); margin-bottom:6px; }
    .login-foot{ text-align:center; font-size:11px; color:var(--muted); margin-top:18px; }
  </style>
</head>
<body>
  <div class="login-shell">
    <div class="login-card">
      <div style="display:flex; justify-content:flex-end; margin-bottom:10px;">{{ theme_btn|safe }}</div>
      <div class="login-hero">
        <div class="login-logo">{{ brand_mark }}</div>
        <div style="text-align:center;">
          <div class="login-h">Create account</div>
          <div class="login-sub">{{ brand_name }}</div>
        </div>
      </div>

      <div class="card" style="padding:22px;">
        {% if error %}
          <div class="msg" style="margin-top:0;">{{ error }}</div>
        {% else %}
          {% if msg %}<div class="msg" style="margin-top:0; margin-bottom:14px;">{{ msg }}</div>{% endif %}
          <form method="post" class="login-fields">
            <input type="hidden" name="token" value="{{ token }}">
            <div>
              <div class="field-label">Email</div>
              <input value="{{ email }}" disabled>
            </div>
            <div>
              <div class="field-label">Name</div>
              <input name="name" value="{{ name }}" placeholder="Your name" required autofocus>
            </div>
            <div>
              <div class="field-label">Password</div>
              <input name="pw" type="password" autocomplete="new-password" placeholder="At least 8 characters" required>
            </div>
            <div>
              <div class="field-label">Confirm password</div>
              <input name="pw2" type="password" autocomplete="new-password" placeholder="Repeat password" required>
            </div>
            <button class="btn-primary" style="width:100%; padding:11px 14px; font-size:14px;">Create account</button>
          </form>
        {% endif %}
      </div>

      <div class="login-foot">
        {% if error %}<a href="{{ url_for('login') }}">Back to sign in</a>{% else %}This invitation link is single-use.{% endif %}
      </div>
    </div>
  </div>
</body></html>
"""

NOCOMPANY_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>No company – Work log</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">
    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div>
          <div class="brand-title">Work log</div>
          <div class="brand-sub">{{ brand_name }}</div>
        </div>
      </div>
      <div class="spacer"></div>
      {% if is_superadmin %}<a class="btn" href="{{ url_for('admin_companies') }}">Admin</a>{% endif %}
      <div class="userchip"><span class="avatar">{{ user_name[:1]|upper }}</span><span>{{ user_name }}</span><span class="role">{{ role_label }}</span></div>
      {{ theme_btn|safe }}
      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>
    <div class="card" style="text-align:center; padding:48px;">
      <div class="card-title" style="justify-content:center;">No company assigned</div>
      {% if is_superadmin %}
        <p class="muted">You have no companies yet. Create one in <a href="{{ url_for('admin_companies') }}">Admin → Companies</a> and start logging.</p>
      {% else %}
        <p class="muted">You are not assigned to any company yet. Please contact an administrator.</p>
      {% endif %}
    </div>
  </div>
</body></html>
"""

ADMIN_COMPANIES_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin – companies</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">
    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div><div class="brand-title">Work log</div><div class="brand-sub">admin – companies</div></div>
      </div>
      <a class="btn" href="{{ url_for('daily') }}">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="19" y1="12" x2="5" y2="12"></line><polyline points="12 19 5 12 12 5"></polyline></svg>
        Back
      </a>
      <a class="btn" href="{{ url_for('admin_audit') }}">Audit</a>
      <div class="spacer"></div>
      <div class="userchip"><span class="avatar">{{ user_name[:1]|upper }}</span><span>{{ user_name }}</span><span class="role">{{ role_label }}</span></div>
      {{ theme_btn|safe }}
      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>

    <div class="card">
      <div class="card-title">Create company</div>
      <form method="post" action="{{ url_for('admin_companies') }}" class="actions" style="align-items:end;">
        <div class="field" style="flex:1;">
          <div class="small">Company name</div>
          <input name="name" placeholder="e.g. Acme Inc." required>
        </div>
        <button class="btn-primary">Create</button>
      </form>
    </div>

    <div class="card">
      <div class="card-title">Companies <span class="muted" style="font-weight:400; font-size:12px; margin-left:auto;">{{ companies_list|length }}</span></div>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Company</th><th class="num" style="width:120px;">Members</th><th style="width:1%;"></th></tr></thead>
          <tbody>
            {% for c in companies_list %}
              <tr>
                <td><b>{{ c.name }}</b></td>
                <td class="num">{{ c.members }}</td>
                <td style="white-space:nowrap;">
                  <a class="btn" href="{{ url_for('company_members', cid=c.id) }}">Manage members</a>
                  <form method="post" action="{{ url_for('company_delete', cid=c.id) }}" style="margin:0; display:inline;" onsubmit="return confirm('Delete this company and ALL its work records? This cannot be undone.');">
                    <button class="btn-danger">Delete</button>
                  </form>
                </td>
              </tr>
            {% endfor %}
            {% if not companies_list %}
              <tr><td colspan="3" class="muted" style="text-align:center; padding:24px;">No companies yet.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</body></html>
"""

COMPANY_MEMBERS_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin – {{ company.name }}</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">
    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div><div class="brand-title">Work log</div><div class="brand-sub">members – {{ company.name }}</div></div>
      </div>
      <a class="btn" href="{{ url_for('admin_companies') }}">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="19" y1="12" x2="5" y2="12"></line><polyline points="12 19 5 12 12 5"></polyline></svg>
        Back
      </a>
      <div class="spacer"></div>
      <div class="userchip"><span class="avatar">{{ user_name[:1]|upper }}</span><span>{{ user_name }}</span><span class="role">{{ role_label }}</span></div>
      {{ theme_btn|safe }}
      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>

    <div class="card">
      <div class="card-title">Invite user</div>
      {% if invite_link %}
        <div class="msg" style="margin-top:0;">
          Invitation link for <b>{{ invite_email }}</b> (send it manually, valid {{ invite_ttl_hours }}h):<br>
          <input readonly value="{{ invite_link }}" onclick="this.select()" style="width:100%; margin-top:8px; font-family:'JetBrains Mono',monospace; font-size:12px;">
        </div>
      {% elif invite_msg %}
        <div class="msg" style="margin-top:0;">{{ invite_msg }}</div>
      {% endif %}
      <form method="post" action="{{ url_for('company_invite', cid=cid) }}" class="actions" style="align-items:end;">
        <div class="field" style="flex:1;">
          <div class="small">Email</div>
          <input name="email" type="email" placeholder="user@example.com" required>
        </div>
        <button class="btn-primary">Create invitation link</button>
      </form>
    </div>

    <div class="card">
      <div class="card-title">Members <span class="muted" style="font-weight:400; font-size:12px; margin-left:auto;">{{ members|length }}</span></div>
      <div class="table-wrap">
        <table>
          <thead><tr><th>User</th><th>Email</th><th style="width:110px;">Status</th><th style="width:150px;">Company role</th><th style="width:1%;"></th></tr></thead>
          <tbody>
            {% for mem in members %}
              <tr>
                <td><b>{{ mem.name or "—" }}</b>{% if mem.is_superadmin %} <span class="role">superadmin</span>{% endif %}</td>
                <td>{{ mem.email }}</td>
                <td>{% if mem.active %}active{% else %}<span class="muted">pending</span>{% endif %}</td>
                <td>
                  <form method="post" action="{{ url_for('company_role_set', cid=cid) }}" style="margin:0; display:flex; gap:6px;">
                    <input type="hidden" name="user_id" value="{{ mem.id }}">
                    <select name="role" onchange="this.form.submit()" style="width:auto;">
                      <option value="member" {% if mem.role == 'member' %}selected{% endif %}>member</option>
                      <option value="admin" {% if mem.role == 'admin' %}selected{% endif %}>manager (read-only)</option>
                    </select>
                  </form>
                </td>
                <td style="white-space:nowrap;">
                  <form method="post" action="{{ url_for('toggle_superadmin', uid=mem.id) }}" style="margin:0; display:inline;">
                    <button class="btn" title="Toggle global superadmin">{% if mem.is_superadmin %}Revoke SA{% else %}Make SA{% endif %}</button>
                  </form>
                  <form method="post" action="{{ url_for('company_remove', cid=cid) }}" style="margin:0; display:inline;" onsubmit="return confirm('Remove this member from the company?');">
                    <input type="hidden" name="user_id" value="{{ mem.id }}">
                    <button class="btn-danger btn-icon" title="Remove from company">
                      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6l-2 14a2 2 0 0 1-2 2H9a2 2 0 0 1-2-2L5 6"></path></svg>
                    </button>
                  </form>
                </td>
              </tr>
            {% endfor %}
            {% if not members %}
              <tr><td colspan="5" class="muted" style="text-align:center; padding:24px;">No members yet. Invite someone above.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</body></html>
"""

ADMIN_AUDIT_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin – audit</title>
  {{ css|safe }}
</head>
<body>
  <div class="container">
    <div class="topbar">
      <div class="brand">
        <span class="brand-mark">{{ brand_mark }}</span>
        <div><div class="brand-title">Work log</div><div class="brand-sub">admin – audit log</div></div>
      </div>
      <a class="btn" href="{{ url_for('admin_companies') }}">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="19" y1="12" x2="5" y2="12"></line><polyline points="12 19 5 12 12 5"></polyline></svg>
        Back
      </a>
      <div class="spacer"></div>
      <span class="userchip"><span class="role">{{ limit }} records</span></span>
      {{ theme_btn|safe }}
      <a class="btn btn-ghost" href="{{ url_for('logout') }}" title="Sign out">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
      </a>
    </div>

    <div class="card" style="padding:0; overflow:hidden;">
      <div class="table-wrap" style="border:none; border-radius:0;">
        <table>
          <thead>
            <tr>
              <th style="width:200px;">Time (UTC)</th>
              <th style="width:160px;">Viewer</th>
              <th style="width:160px;">Target</th>
              <th style="width:140px;">Company</th>
              <th style="width:120px;">Action</th>
              <th style="width:100px;">Month</th>
              <th>IP</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
              <tr>
                <td class="num" style="font-variant-numeric:tabular-nums; color:var(--muted); font-size:12.5px;">{{ r.ts_utc }}</td>
                <td><b>{{ r.viewer_email }}</b></td>
                <td>{{ r.target_email }}</td>
                <td>{{ r.company_name }}</td>
                <td><span style="display:inline-flex; padding:3px 8px; background:var(--accent-soft); color:var(--accent); border-radius:6px; font-size:11px; font-weight:600; text-transform:uppercase; letter-spacing:.04em;">{{ r.action }}</span></td>
                <td class="num">{{ r.month }}</td>
                <td style="font-family:'JetBrains Mono',monospace; font-size:12px; color:var(--muted);">{{ r.ip }}</td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="7" class="muted" style="text-align:center; padding:24px;">No records.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</body></html>
"""


# ----------------------------
# Render helpers
# ----------------------------
def role_badge(u: "UserCtx", role: Optional[str]) -> str:
    if u.is_superadmin:
        return "superadmin"
    if role == "admin":
        return "manager"
    return "member"

def base_ctx(u: "UserCtx", company_id=None, role=None) -> Dict[str, Any]:
    return dict(
        css=BASE_CSS,
        theme_btn=THEME_BTN,
        brand_name=BRAND_NAME,
        brand_mark=BRAND_MARK,
        companies=companies_for_user(u),
        company_id=company_id,
        user_name=u.name,
        role_label=role_badge(u, role),
        is_superadmin=u.is_superadmin,
    )

def safe_name(s: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in (s or "")).strip("_") or "user"


# ----------------------------
# Routes: auth
# ----------------------------
@APP.get("/login")
def login():
    init_db()
    return render_template_string(LOGIN_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                  brand_name=BRAND_NAME, brand_mark=BRAND_MARK, msg="")

@APP.post("/login")
def login_post():
    init_db()
    email = (request.form.get("email") or "").strip().lower()
    pw = request.form.get("pw") or ""
    nxt = request.args.get("next") or url_for("daily")

    def fail(msg):
        return render_template_string(LOGIN_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                      brand_name=BRAND_NAME, brand_mark=BRAND_MARK, msg=msg)

    if not email or not pw:
        return fail("Missing email or password.")

    user = get_user_by_email(email)
    if not user or not user["active"] or not user["password_hash"]:
        return fail("Invalid email or password.")
    if not check_password_hash(user["password_hash"], pw):
        return fail("Invalid email or password.")

    session.clear()
    session["user_id"] = int(user["id"])
    u = current_user()
    if u:
        comps = companies_for_user(u)
        if comps:
            session["company_id"] = int(comps[0]["id"])
    return redirect(nxt)

@APP.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


ACCOUNT_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Change password - Work log</title>
  {{ css|safe }}
  <style>
    .login-shell{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:24px; }
    .login-card{ width:min(440px, 100%); }
    .login-fields{ display:flex; flex-direction:column; gap:14px; }
    .field-label{ font-size:12px; font-weight:600; color:var(--fg-soft); margin-bottom:6px; }
  </style>
</head>
<body>
  <div class="login-shell">
    <div class="login-card">
      <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:14px;">
        <a class="btn" href="{{ url_for('daily') }}">Back</a>
        {{ theme_btn|safe }}
      </div>
      <div class="card" style="padding:22px;">
        <div class="card-title" style="margin-bottom:14px;">Change password</div>
        {% if msg %}<div class="msg" style="margin-top:0; margin-bottom:14px;">{{ msg }}</div>{% endif %}
        {% if ok %}<div class="msg" style="margin-top:0; margin-bottom:14px; color:var(--accent);">Password changed successfully.</div>{% endif %}
        <form method="post" class="login-fields">
          <div>
            <div class="field-label">Current password</div>
            <input name="current" type="password" autocomplete="current-password" required autofocus>
          </div>
          <div>
            <div class="field-label">New password (min 8 characters)</div>
            <input name="new" type="password" autocomplete="new-password" required>
          </div>
          <div>
            <div class="field-label">Confirm new password</div>
            <input name="confirm" type="password" autocomplete="new-password" required>
          </div>
          <button class="btn-primary" style="width:100%; padding:11px 14px; font-size:14px;">Change password</button>
        </form>
      </div>
      <div style="text-align:center; font-size:11px; color:var(--muted); margin-top:18px;">{{ user_name }} &middot; {{ brand_name }}</div>
    </div>
  </div>
</body>
</html>
"""

@APP.get("/account")
def account():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    return render_template_string(ACCOUNT_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                  brand_name=BRAND_NAME, user_name=u.name, msg="", ok=False)

@APP.post("/account")
def account_post():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()

    def render(msg="", ok=False):
        return render_template_string(ACCOUNT_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                      brand_name=BRAND_NAME, user_name=u.name, msg=msg, ok=ok)

    cur = request.form.get("current") or ""
    new = request.form.get("new") or ""
    confirm = request.form.get("confirm") or ""
    row = get_user_by_id(u.id)
    if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], cur):
        return render("Current password is incorrect.")
    if len(new) < 8:
        return render("New password must be at least 8 characters.")
    if new != confirm:
        return render("New passwords do not match.")
    db = get_db()
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new), u.id))
    db.commit()
    return render(ok=True)

@APP.route("/register", methods=["GET", "POST"])
def register():
    init_db()
    token = (request.values.get("token") or "").strip()

    def err(message):
        return render_template_string(REGISTER_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                      brand_name=BRAND_NAME, brand_mark=BRAND_MARK,
                                      error=message, email="", name="", token="", msg="")

    if not token:
        return err("Invalid or missing invitation link.")

    row = get_db().execute(
        "SELECT * FROM users WHERE invite_token=? AND active=0", (token,)
    ).fetchone()
    if not row:
        return err("This invitation link is invalid or has already been used.")
    if (row["invite_expires"] or "") and row["invite_expires"] < now_iso():
        return err("This invitation link has expired. Ask an administrator for a new one.")

    if request.method == "GET":
        return render_template_string(REGISTER_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                      brand_name=BRAND_NAME, brand_mark=BRAND_MARK,
                                      error="", email=row["email"], name=(row["name"] or ""),
                                      token=token, msg="")

    name = (request.form.get("name") or "").strip()
    pw = request.form.get("pw") or ""
    pw2 = request.form.get("pw2") or ""

    def form_msg(message):
        return render_template_string(REGISTER_TEMPLATE, css=BASE_CSS, theme_btn=THEME_BTN,
                                      brand_name=BRAND_NAME, brand_mark=BRAND_MARK,
                                      error="", email=row["email"], name=name, token=token, msg=message)

    if len(pw) < 8:
        return form_msg("Password must be at least 8 characters.")
    if pw != pw2:
        return form_msg("Passwords do not match.")

    db = get_db()
    db.execute(
        "UPDATE users SET name=?, password_hash=?, active=1, invite_token=NULL, invite_expires=NULL WHERE id=?",
        (name or row["email"], generate_password_hash(pw), int(row["id"])),
    )
    db.commit()

    session.clear()
    session["user_id"] = int(row["id"])
    u = current_user()
    if u:
        comps = companies_for_user(u)
        if comps:
            session["company_id"] = int(comps[0]["id"])
    return redirect(url_for("daily"))

@APP.route("/switch_company", methods=["GET", "POST"])
def switch_company():
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    cid = request.values.get("company_id")
    try:
        cid = int(cid)
    except (TypeError, ValueError):
        cid = None
    if cid is not None and company_role(u, cid):
        session["company_id"] = cid
    return redirect(request.referrer or url_for("daily"))


# ----------------------------
# Routes: daily
# ----------------------------
@APP.get("/")
def daily():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return render_template_string(NOCOMPANY_TEMPLATE, **base_ctx(u))

    d_str = request.args.get("d") or date.today().isoformat()
    msg = request.args.get("_msg", "")
    try:
        d = parse_date(d_str)
    except Exception:
        d = date.today()
        msg = "Invalid date format."

    day_row_full = get_day_row(d, u.id, company_id)
    if day_row_full:
        day_id = int(day_row_full["id"])
        work_text = day_row_full["work_text"] or ""
        blocks = fetch_blocks(day_id, u.id, company_id)
        day_total = day_total_hours(day_id, u.id, company_id)
    else:
        work_text = ""
        blocks = []
        day_total = 0.0

    m = d.strftime("%Y-%m")
    month_total = month_total_hours(m, u.id, company_id)

    default_from = "06:00"
    default_to = "15:00"
    if blocks:
        last_to = blocks[-1]["t_to"]
        default_from = last_to
        try:
            lt = parse_time(last_to)
            mm = minutes_of(lt) + 60
            if mm > 23 * 60 + 59:
                mm = 23 * 60 + 59
            default_to = f"{mm//60:02d}:{mm%60:02d}"
        except Exception:
            pass

    if msg == "overlap":
        msg = "This block overlaps an existing block – not added."
    elif msg == "badtime":
        msg = "Invalid block time (To must be greater than From)."

    ctx = base_ctx(u, company_id, role)
    ctx.update(dict(
        readonly=False,
        d_iso=d.isoformat(),
        d_human=fmt_date_ddmmyyyy(d),
        work_text=work_text,
        blocks=blocks,
        day_total=day_total,
        month_total=month_total,
        default_from=default_from,
        default_to=default_to,
        m=m,
        msg=msg or "",
    ))
    return render_template_string(DAILY_TEMPLATE, **ctx)

@APP.get("/today")
def today():
    gate = require_login()
    if gate:
        return gate
    return redirect(url_for("daily", d=date.today().isoformat()))


# ----------------------------
# Routes: daily write actions
# ----------------------------
def _company_or_none(u):
    return current_company(u)

@APP.post("/save_day")
def save_day():
    init_db()
    gate = require_login()
    if gate:
        if request.headers.get("X-Requested-With") == "fetch":
            return ("login required", 401)
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        if request.headers.get("X-Requested-With") == "fetch":
            return ("no company", 400)
        return redirect(url_for("daily"))

    d = parse_date(request.form["work_date"])
    day_id = ensure_day(d, u.id, company_id)
    text = request.form.get("work_text", "")

    db = get_db()
    db.execute("UPDATE days SET work_text=? WHERE user_id=? AND company_id=? AND id=?",
               (text, u.id, company_id, day_id))
    db.commit()

    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify(ok=True, saved_at=datetime.now().strftime("%H:%M:%S"))
    return redirect(url_for("daily", d=d.isoformat()))

@APP.post("/copy_yesterday_text")
def copy_yesterday_text():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return redirect(url_for("daily"))

    d = parse_date(request.form["work_date"])
    day_id = ensure_day(d, u.id, company_id)
    db = get_db()

    today_text = (db.execute("SELECT work_text FROM days WHERE user_id=? AND company_id=? AND id=?",
                             (u.id, company_id, day_id)).fetchone()["work_text"] or "").strip()
    y = d - timedelta(days=1)
    y_row = db.execute("SELECT work_text FROM days WHERE user_id=? AND company_id=? AND work_date=?",
                       (u.id, company_id, y.isoformat())).fetchone()
    y_text = ((y_row["work_text"] if y_row else "") or "").strip()

    if not y_text:
        return redirect(url_for("daily", d=d.isoformat()))

    new_text = y_text if not today_text else (today_text + "\n\n---\n(yesterday's text)\n" + y_text)
    db.execute("UPDATE days SET work_text=? WHERE user_id=? AND company_id=? AND id=?",
               (new_text, u.id, company_id, day_id))
    db.commit()
    return redirect(url_for("daily", d=d.isoformat()))

@APP.post("/copy_yesterday_blocks")
def copy_yesterday_blocks():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return redirect(url_for("daily"))

    d = parse_date(request.form["work_date"])
    today_id = ensure_day(d, u.id, company_id)

    y = d - timedelta(days=1)
    y_day = get_day_row(y, u.id, company_id)
    if not y_day:
        return redirect(url_for("daily", d=d.isoformat()))
    y_id = int(y_day["id"])

    db = get_db()
    today_intervals = fetch_blocks_intervals(today_id, u.id, company_id)
    y_blocks = db.execute(
        "SELECT t_from, t_to, hours FROM blocks WHERE user_id=? AND company_id=? AND day_id=? ORDER BY t_from",
        (u.id, company_id, y_id),
    ).fetchall()

    for b in y_blocks:
        tf = parse_time(b["t_from"])
        tt = parse_time(b["t_to"])
        nf, nt = minutes_of(tf), minutes_of(tt)
        if any(intervals_overlap(nf, nt, ef, et) for (ef, et) in today_intervals):
            continue
        db.execute(
            "INSERT INTO blocks(user_id, company_id, day_id, t_from, t_to, hours) VALUES(?,?,?,?,?,?)",
            (u.id, company_id, today_id, b["t_from"], b["t_to"], float(b["hours"])),
        )
        today_intervals.append((nf, nt))

    db.commit()
    return redirect(url_for("daily", d=d.isoformat()))

def _add_time_block(preset: bool):
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return redirect(url_for("daily"))
    d = parse_date(request.form["work_date"])
    day_id = ensure_day(d, u.id, company_id)

    t_from = parse_time(request.form["t_from"])
    t_to = parse_time(request.form["t_to"])
    nf, nt = minutes_of(t_from), minutes_of(t_to)

    if nt <= nf:
        return redirect(url_for("daily", d=d.isoformat(), _msg="badtime"))

    existing = fetch_blocks_intervals(day_id, u.id, company_id)
    if any(intervals_overlap(nf, nt, ef, et) for (ef, et) in existing):
        return redirect(url_for("daily", d=d.isoformat(), _msg="overlap"))

    h = hours_between(d, t_from, t_to)
    db = get_db()
    db.execute(
        "INSERT INTO blocks(user_id, company_id, day_id, t_from, t_to, hours) VALUES(?,?,?,?,?,?)",
        (u.id, company_id, day_id, t_from.strftime("%H:%M"), t_to.strftime("%H:%M"), h),
    )
    db.commit()
    return redirect(url_for("daily", d=d.isoformat()))

@APP.post("/add_block")
def add_block():
    init_db()
    gate = require_login()
    if gate:
        return gate
    return _add_time_block(preset=False)

@APP.post("/add_preset")
def add_preset():
    init_db()
    gate = require_login()
    if gate:
        return gate
    return _add_time_block(preset=True)

@APP.post("/delete_block")
def delete_block():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return redirect(url_for("daily"))
    d = parse_date(request.form["work_date"])
    block_id = int(request.form["block_id"])
    db = get_db()
    db.execute("DELETE FROM blocks WHERE user_id=? AND company_id=? AND id=?", (u.id, company_id, block_id))
    db.commit()
    return redirect(url_for("daily", d=d.isoformat()))


# ----------------------------
# Routes: month
# ----------------------------
@APP.get("/month")
def month_view():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return render_template_string(NOCOMPANY_TEMPLATE, **base_ctx(u))

    can_view_others = is_company_admin(u, company_id)
    target_id = effective_target_id(u, company_id, request.args.get("u"))

    db = get_db()
    months = [r["m"] for r in db.execute(
        "SELECT DISTINCT substr(work_date,1,7) AS m FROM days WHERE user_id=? AND company_id=? ORDER BY m DESC",
        (target_id, company_id),
    ).fetchall()]
    if not months:
        months = [date.today().strftime("%Y-%m")]

    sel = request.args.get("m") or months[0]
    if sel not in months:
        sel = months[0]

    if target_id != u.id:
        audit_log(company_id, u.id, target_id, "month_view", sel)

    rows = rows_for_month(sel, target_id, company_id)
    mtotal = month_total_hours(sel, target_id, company_id)
    can_delete_day = (target_id == u.id)
    members = company_users(company_id) if can_view_others else []

    ctx = base_ctx(u, company_id, role)
    ctx.update(dict(
        target_id=target_id,
        target_name=user_display(target_id),
        is_viewing_other=(target_id != u.id),
        months=months,
        sel=sel,
        rows=rows,
        month_total=mtotal,
        can_delete_day=can_delete_day,
        members=members,
        can_view_others=can_view_others,
    ))
    return render_template_string(MONTH_TEMPLATE, **ctx)

@APP.post("/delete_day")
def delete_day():
    init_db()
    gate = require_login()
    if gate:
        return gate
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return redirect(url_for("daily"))
    day_id = int(request.form["day_id"])
    m = request.form.get("m") or date.today().strftime("%Y-%m")
    delete_day_by_id(day_id=day_id, user_id=u.id, company_id=company_id)
    return redirect(url_for("month_view", m=m))


# ----------------------------
# Routes: exports
# ----------------------------
def _export_target():
    """Return (u, company_id, target_id, month) for an export, or a redirect response."""
    u = current_user()
    company_id, role = current_company(u)
    if company_id is None:
        return None
    target_id = effective_target_id(u, company_id, request.args.get("u"))
    m = request.args.get("m") or date.today().strftime("%Y-%m")
    return u, company_id, target_id, m

@APP.get("/export.csv")
def export_csv():
    init_db()
    gate = require_login()
    if gate:
        return gate
    tgt = _export_target()
    if tgt is None:
        return redirect(url_for("daily"))
    u, company_id, target_id, m = tgt
    if target_id != u.id:
        audit_log(company_id, u.id, target_id, "export_csv", m)

    display = user_display(target_id)
    rows = rows_for_month(m, target_id, company_id)
    total = month_total_hours(m, target_id, company_id)

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["user", "date", "from", "to", "hours", "work", "day_total"])
    for r in rows:
        w.writerow([display, r["work_date"], r["t_from"], r["t_to"], r["hours"], r["work_text"], r["day_total"]])
    w.writerow([])
    w.writerow(["MONTH TOTAL", "", "", "", "", "", f"{total:.1f}"])

    out = io.BytesIO(buf.getvalue().encode("utf-8"))
    out.seek(0)
    return send_file(out, mimetype="text/csv", as_attachment=True,
                     download_name=f"worklog_{safe_name(display)}_{m}.csv")

@APP.get("/export.md")
def export_md():
    init_db()
    gate = require_login()
    if gate:
        return gate
    tgt = _export_target()
    if tgt is None:
        return redirect(url_for("daily"))
    u, company_id, target_id, m = tgt
    if target_id != u.id:
        audit_log(company_id, u.id, target_id, "export_md", m)

    display = user_display(target_id)
    rows = rows_for_month(m, target_id, company_id)
    total = month_total_hours(m, target_id, company_id)

    out_lines = []
    out_lines.append("# Worklog")
    out_lines.append("")
    out_lines.append(f"- **User:** {md_escape(display)}")
    out_lines.append(f"- **Month:** {md_escape(m)}")
    out_lines.append("")
    out_lines.append("| Date | From | To | Hours | Work | Day total |")
    out_lines.append("|---|---:|---:|---:|---|---:|")
    for r in rows:
        out_lines.append(
            f"| {md_escape(r.get('work_date',''))} | {md_escape(r.get('t_from',''))} | {md_escape(r.get('t_to',''))} | "
            f"{md_escape(r.get('hours',''))} | {md_escape(r.get('work_text',''))} | {md_escape(r.get('day_total',''))} |"
        )
    out_lines.append("")
    out_lines.append(f"**MONTH TOTAL:** {total:.1f} h")
    out_lines.append("")

    data = "\n".join(out_lines).encode("utf-8")
    bio = io.BytesIO(data)
    bio.seek(0)
    return send_file(bio, mimetype="text/markdown; charset=utf-8", as_attachment=True,
                     download_name=f"worklog_{safe_name(display)}_{m}.md")

@APP.get("/export.xlsx")
def export_xlsx():
    init_db()
    gate = require_login()
    if gate:
        return gate
    tgt = _export_target()
    if tgt is None:
        return redirect(url_for("daily"))
    u, company_id, target_id, m = tgt
    if target_id != u.id:
        audit_log(company_id, u.id, target_id, "export_xlsx", m)

    display = user_display(target_id)
    rows = rows_for_month(m, target_id, company_id)
    total = month_total_hours(m, target_id, company_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Worklog"
    ws.append(["user", "date", "from", "to", "hours", "work", "day_total"])
    for r in rows:
        ws.append([display, r["work_date"], r["t_from"], r["t_to"], r["hours"], r["work_text"], r["day_total"]])
    ws.append([])
    ws.append(["MONTH TOTAL", "", "", "", "", "", float(f"{total:.1f}")])

    widths = [22, 12, 8, 8, 8, 70, 12]
    for i, w_ in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"worklog_{safe_name(display)}_{m}.xlsx")


# ----------------------------
# Routes: admin (superadmin only)
# ----------------------------
@APP.get("/admin")
def admin_home():
    gate = require_superadmin()
    if gate:
        return gate
    return redirect(url_for("admin_companies"))

@APP.route("/admin/companies", methods=["GET", "POST"])
def admin_companies():
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    u = current_user()
    db = get_db()
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if name:
            db.execute("INSERT INTO companies(name, created) VALUES(?,?)", (name, now_iso()))
            db.commit()
        return redirect(url_for("admin_companies"))

    comps = db.execute(
        "SELECT c.id, c.name, "
        "(SELECT COUNT(*) FROM company_members m WHERE m.company_id=c.id) AS members "
        "FROM companies c ORDER BY c.name"
    ).fetchall()
    company_id, role = current_company(u)
    ctx = base_ctx(u, company_id, role)
    ctx.update(dict(companies_list=comps))
    return render_template_string(ADMIN_COMPANIES_TEMPLATE, **ctx)

@APP.post("/admin/company/<int:cid>/delete")
def company_delete(cid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    db = get_db()
    # days/blocks/audit have company_id but no FK cascade -> remove explicitly.
    db.execute("DELETE FROM blocks WHERE company_id=?", (cid,))
    db.execute("DELETE FROM days WHERE company_id=?", (cid,))
    db.execute("DELETE FROM company_members WHERE company_id=?", (cid,))
    db.execute("DELETE FROM companies WHERE id=?", (cid,))
    db.commit()
    if session.get("company_id") == cid:
        session.pop("company_id", None)
    return redirect(url_for("admin_companies"))

@APP.get("/admin/company/<int:cid>/members")
def company_members(cid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    u = current_user()
    db = get_db()
    company = db.execute("SELECT id, name FROM companies WHERE id=?", (cid,)).fetchone()
    if not company:
        abort(404)
    members = company_users(cid)
    company_id, role = current_company(u)
    ctx = base_ctx(u, company_id, role)
    ctx.update(dict(
        company=company,
        members=members,
        cid=cid,
        invite_link=session.pop("invite_link", None),
        invite_email=session.pop("invite_email", None),
        invite_msg=session.pop("invite_msg", None),
        invite_ttl_hours=INVITE_TTL_HOURS,
    ))
    return render_template_string(COMPANY_MEMBERS_TEMPLATE, **ctx)

@APP.post("/admin/company/<int:cid>/invite")
def company_invite(cid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    db = get_db()
    company = db.execute("SELECT id FROM companies WHERE id=?", (cid,)).fetchone()
    if not company:
        abort(404)
    email = (request.form.get("email") or "").strip().lower()
    if not email:
        session["invite_msg"] = "Email is required."
        return redirect(url_for("company_members", cid=cid))

    existing = get_user_by_email(email)
    if existing:
        db.execute("INSERT OR IGNORE INTO company_members(company_id, user_id, role) VALUES(?,?, 'member')",
                   (cid, int(existing["id"])))
        if not existing["active"]:
            token = secrets.token_urlsafe(32)
            expires = (datetime.utcnow() + timedelta(hours=INVITE_TTL_HOURS)).replace(microsecond=0).isoformat() + "Z"
            db.execute("UPDATE users SET invite_token=?, invite_expires=? WHERE id=?",
                       (token, expires, int(existing["id"])))
            db.commit()
            session["invite_link"] = request.host_url.rstrip("/") + url_for("register", token=token)
            session["invite_email"] = email
        else:
            db.commit()
            session["invite_msg"] = f"User {email} already exists and was added to the company."
        return redirect(url_for("company_members", cid=cid))

    token = secrets.token_urlsafe(32)
    expires = (datetime.utcnow() + timedelta(hours=INVITE_TTL_HOURS)).replace(microsecond=0).isoformat() + "Z"
    cur = db.execute(
        "INSERT INTO users(email, name, password_hash, is_superadmin, active, invite_token, invite_expires, created) "
        "VALUES(?,?,?,0,0,?,?,?)",
        (email, "", None, token, expires, now_iso()),
    )
    new_id = int(cur.lastrowid)
    db.execute("INSERT OR IGNORE INTO company_members(company_id, user_id, role) VALUES(?,?, 'member')",
               (cid, new_id))
    db.commit()
    session["invite_link"] = request.host_url.rstrip("/") + url_for("register", token=token)
    session["invite_email"] = email
    return redirect(url_for("company_members", cid=cid))

@APP.post("/admin/company/<int:cid>/role")
def company_role_set(cid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    role = request.form.get("role")
    if role not in ("member", "admin"):
        role = "member"
    try:
        uid = int(request.form.get("user_id"))
    except (TypeError, ValueError):
        return redirect(url_for("company_members", cid=cid))
    db = get_db()
    db.execute("UPDATE company_members SET role=? WHERE company_id=? AND user_id=?", (role, cid, uid))
    db.commit()
    return redirect(url_for("company_members", cid=cid))

@APP.post("/admin/company/<int:cid>/remove")
def company_remove(cid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    try:
        uid = int(request.form.get("user_id"))
    except (TypeError, ValueError):
        return redirect(url_for("company_members", cid=cid))
    db = get_db()
    db.execute("DELETE FROM company_members WHERE company_id=? AND user_id=?", (cid, uid))
    db.commit()
    return redirect(url_for("company_members", cid=cid))

@APP.post("/admin/user/<int:uid>/superadmin")
def toggle_superadmin(uid):
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    db = get_db()
    row = db.execute("SELECT is_superadmin FROM users WHERE id=?", (uid,)).fetchone()
    if row:
        if row["is_superadmin"]:
            # Do not allow removing the last active superadmin.
            cnt = db.execute("SELECT COUNT(*) AS c FROM users WHERE is_superadmin=1 AND active=1").fetchone()["c"]
            if int(cnt) > 1:
                db.execute("UPDATE users SET is_superadmin=0 WHERE id=?", (uid,))
        else:
            db.execute("UPDATE users SET is_superadmin=1 WHERE id=?", (uid,))
        db.commit()
    # Redirect back to referrer (members page) if available.
    return redirect(request.referrer or url_for("admin_companies"))

@APP.get("/admin/audit")
def admin_audit():
    init_db()
    gate = require_superadmin()
    if gate:
        return gate
    u = current_user()
    limit = 500
    rows = list_audit(limit=limit)
    company_id, role = current_company(u)
    ctx = base_ctx(u, company_id, role)
    ctx.update(dict(rows=rows, limit=limit))
    return render_template_string(ADMIN_AUDIT_TEMPLATE, **ctx)


if __name__ == "__main__":
    init_db()
    APP.run(host="0.0.0.0", port=5000, debug=False)
